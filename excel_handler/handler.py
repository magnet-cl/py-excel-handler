""" This document defines the excel_handler module """
from __future__ import print_function, absolute_import
from builtins import str, object
import xlsxwriter
import datetime
from .fields import Field

from collections import namedtuple
from future.utils import with_metaclass

from openpyxl.utils.datetime import from_excel
from openpyxl import load_workbook


class FieldNotFound(Exception):
    pass


class ReapeatedColumn(Exception):
    pass


RowError = namedtuple("RowError", "row, row_data, error, field_name")


class ExcelHandlerMetaClass(type):
    def __new__(cls, name, bases, attrs):
        fieldname_to_field = {}

        for base in bases[::-1]:
            if hasattr(base, "fieldname_to_field"):
                fieldname_to_field.update(base.fieldname_to_field)

        cols = {}

        for k, v in list(attrs.items()):
            if isinstance(v, Field):
                field = attrs.pop(k)
                field.name = k
                if field.verbose_name == "":
                    field.verbose_name = field.name
                if field.col < 0:
                    field._distance_from_last = field.col

                if field.col in cols:
                    raise ReapeatedColumn(
                        "{} collides with field {} on column {}".format(
                            field.name, cols[field.col].name, field.col
                        )
                    )

                cols[field.col] = field

                fieldname_to_field[k] = field

        attrs["fieldname_to_field"] = fieldname_to_field
        attrs["fields"] = sorted(
            list(fieldname_to_field.values()), key=lambda field: field.col
        )

        sup = super(ExcelHandlerMetaClass, cls)

        this = sup.__new__(cls, name, bases, attrs)
        field_count = len(fieldname_to_field)

        for field_name, field in list(fieldname_to_field.items()):
            try:
                if field._distance_from_last < 0:
                    field.col = field_count + field._distance_from_last
            except:
                pass

        return this


class ExcelHandler(with_metaclass(ExcelHandlerMetaClass, object)):
    """ExcelHandler is a class that is used to wrap common operations in
    excel files"""

    def __init__(self, path=None, excel_file=None, mode="r", on_demand=False):
        if path is None and excel_file is None:
            raise Exception("path or excel_file requried")
        if path is not None and excel_file is not None:
            raise Exception("Only specify path or excel_file, not both")
        if mode == "r":
            if path:
                self.workbook = load_workbook(
                    filename=path,
                )
            else:
                self.workbook = load_workbook(
                    filename=excel_file,
                )
            self.sheet = self.workbook.worksheets[0]

        else:
            self.path = path

            self.workbook = xlsxwriter.Workbook(self.path, {"nan_inf_to_errors": True})

            self.set_default_formats()

        self.parser = None

    def set_default_formats(self):
        self.date_format = self.workbook.add_format({"num_format": "YYYY-MM-DD"})
        self.datetime_format = self.workbook.add_format(
            {"num_format": "YYYY-MM-DD HH:MM:SS"}
        )
        self.time_format = self.workbook.add_format({"num_format": "HH:MM:SS"})

    def set_row_formats_from_example(self, row):
        i = 0
        for cel in row:
            if isinstance(cel, datetime.date):
                self.sheet.set_column(i, i, 18, cell_format=self.date_format)
            elif isinstance(cel, datetime.datetime):
                self.sheet.set_column(i, i, 18, cell_format=self.datetime_format)
            elif isinstance(cel, datetime.time):
                self.sheet.set_column(i, i, 18, cell_format=self.time_format)
            i += 1

    def add_sheet(self, name):

        self.sheet = self.workbook.add_worksheet(name)

    def set_sheet(self, sheet_index):
        """sets the current sheet with the given sheet_index"""
        self.sheet = self.workbook.worksheets[sheet_index]

    def set_sheet_by_name(self, sheet_name):
        """sets the current sheet with the given sheet name"""
        self.sheet = self.workbook[sheet_name]

    def parse_date(self, value):
        return from_excel(value).date()

    def read_rows(self, column_structure, starting_row=1, max_rows=None):
        """Reads the current sheet from the starting row to the last row or up
        to a max of max_rows if greater than 0

        returns an array with the data

        """
        data = []
        rows = self.sheet.iter_rows(
            min_row=starting_row,
            max_row=max_rows,
            max_col=len(column_structure),
        )

        for row in rows:
            column_data = {}
            for cell in row:
                value = cell.value
                column_name = list(column_structure)[cell.col_idx - 1]
                column_data[column_name] = value
            data.append(column_data)

        return data

    def _read(
        self,
        skip_titles=False,
        failfast=False,
        ignore_blank_rows=True,
        include_rowx=False,
        return_errors=False,
        starting_row=0,
    ):
        """
        Using the structure defined with the Field attributes, reads the excel
        and returns the data in an array of dicts
        """
        data = []
        errors = []
        row = starting_row
        if skip_titles:
            row += 1

        # prepare the read for each field
        for field in self.fields:
            field.prepare_read()

        while True:
            row_data = {}
            data_read = False
            continue_while = False
            blank_row = True

            for field in self.fields:
                field_name = field.name
                try:
                    value = self.sheet.cell(colx=field.col, rowx=row).value
                except:
                    if hasattr(field, "default"):
                        row_data[field.name] = field.default
                else:
                    if value != "":
                        blank_row = False

                    try:
                        row_data[field.name] = field.cast(
                            value,
                            self.workbook,
                            row_data,
                        )
                    except Exception as err:
                        if not err.args:
                            err.args = ("",)
                        msg = 'Cannot read row "{}" : Column {}, {}'.format(
                            row + 1, str(field.verbose_name), err.args[0]
                        )
                        err.args = (msg,) + err.args[1:]
                        if failfast:
                            raise
                        else:
                            row_data[field.name] = value
                            if return_errors:
                                errors.append(
                                    RowError(
                                        row=row,
                                        row_data=row_data,
                                        error=msg,
                                        field_name=field_name,
                                    )
                                )
                            else:
                                print(msg)
                            continue_while = True
                        break

                    data_read = True

                if include_rowx:
                    row_data["rowx"] = row

            row += 1

            if continue_while:
                continue

            if not data_read:
                if return_errors:
                    return data, errors
                return data

            if not blank_row or not ignore_blank_rows:
                data.append(row_data)

        if return_errors:
            return data, errors
        return data

    def read(
        self,
        skip_titles=False,
        failfast=False,
        ignore_blank_rows=True,
        include_rowx=False,
        return_errors=False,
        starting_row=1,
    ):
        """
        Using the structure defined with the Field attributes, reads the excel
        and returns the data in an array of dicts
        """
        data = []
        errors = []

        min_row = 1
        if skip_titles:
            min_row += 1

        if not starting_row == 1:
            min_row = starting_row

        # prepare the read for each field
        for field in self.fields:
            field.prepare_read()

        for row in self.sheet.iter_rows(min_row=min_row):
            row_data = {}
            empty_fields = []
            has_errors = False

            for cell in row:
                value = cell.value

                try:
                    # get fields by column
                    field = self.fields[cell.column - 1]
                except Exception:
                    break

                if value is None:
                    empty_fields.append(value)

                if value is None and hasattr(field, "default"):
                    default_value = field.default
                    if callable(default_value):
                        value = default_value()
                    else:
                        value = default_value
                else:
                    try:
                        value = field.cast(
                            value,
                            self.workbook,
                            row_data,
                        )
                    except Exception as err:
                        has_errors = True
                        if failfast:
                            raise
                        if return_errors:
                            msg = f'Cannot read row "{row.id}" : Column {str(field.verbose_name)}, {err.args[0]}'
                            errors.append(
                                RowError(
                                    row=row,
                                    row_data=row_data,
                                    error=msg,
                                    field_name=field.name,
                                )
                            )
                        break

                row_data[field.name] = value

            if has_errors:
                continue

            if ignore_blank_rows:
                if not len(empty_fields) == len(row_data):
                    data.append(row_data)
            else:
                data.append(row_data)

        if return_errors:
            return data, errors
        return data

    def save(self):
        """Save document"""

        # xlwt save
        # self.workbook.save(self.path)
        self.workbook.close()

    def set_title_format(self, formt):
        pass

    def set_row_format(self):
        return None

    def write_rows(self, rows, col_offset=0, row_offset=0, set_titles=False):
        """Write rows in the current sheet"""

        title_formt = self.workbook.add_format()
        row_formt = self.set_row_format()

        if set_titles:
            self.set_title_format(title_formt)
        else:
            title_formt = row_formt

        for y, row in enumerate(rows):
            # set titles
            if y == 0:
                formt = title_formt
            else:
                formt = row_formt

            row_y = row_offset + y

            for x, value in enumerate(row):
                row_x = col_offset + x

                self.sheet.write(row_y, row_x, value, formt)

    def write_columns(self, columns, row_offset=0, col_offset=0, set_titles=False):
        """Write columns in the current sheet"""

        if set_titles:
            formt = self.workbook.add_format()
            self.set_title_format(formt)
        else:
            formt = None

        for x, column in enumerate(columns):
            # set titles
            if x > 0:
                formt = None

            column_x = col_offset + x

            for y, value in enumerate(column):
                column_y = row_offset + y
                self.sheet.write(column_y, column_x, value, formt)

    def write(self, data, set_titles=False):
        row = 0

        # set titles
        if set_titles:
            formt = self.workbook.add_format()
            self.set_title_format(formt)

            for field_name, field in list(self.fieldname_to_field.items()):
                self.sheet.write(0, field.col, str(field.verbose_name), formt)
            row = 1

        # set format and prepare the write for each field
        for field_name, field in self.fieldname_to_field.items():
            field.set_column_format(self)
            field.prepare_write()

        for row_data in data:
            for field_name, value in row_data.items():
                try:
                    field = self.fieldname_to_field[field_name]
                except KeyError:
                    pass
                else:
                    field.write(self.workbook, self.sheet, row, value)
            row += 1
